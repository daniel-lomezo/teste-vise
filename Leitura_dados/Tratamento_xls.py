# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from Leitura_dados.Busca_cep import Requesicao_Cep
from pprint import pprint
import os

class Tratamento_arquivo(object):

    def Visualizar_dados(self):
        arquivo_excel = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dados-pessoais.xlsx'))
        planilha1 = arquivo_excel['Página1']
        for i in range(1, 51 + 2):
            for j in range(1, 6 + 1):
                if planilha1.cell(row=i, column=j).value != None and planilha1.cell(row=i, column=j).value != '':
                    print(planilha1.cell(row=i, column=j).value, end=" - ")
            if planilha1.cell(row=i, column=j).value != None and planilha1.cell(row=i, column=j).value != '':
                cep = Requesicao_Cep.Pesquisa_Cep('0', planilha1.cell(row=i, column=5).value)
                pprint('Endereço completo {} '.format(cep))
            print('\n', '\n')



