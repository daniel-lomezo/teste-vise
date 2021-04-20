# -*- coding: utf-8 -*-
import pymysql.cursors
import os
import csv
from Leitura_dados.Tradutor import Google_Trad
from Leitura_dados.Conversor_moeda import Busca_moedas
from openpyxl import load_workbook
from Leitura_dados.Busca_cep import Requesicao_Cep
from pprint import pprint
import os

Create_table = """

CREATE TABLE Desafio (
id_usuario INT(6) UNSIGNED AUTO_INCREMENT PRIMARY KEY,
nome VARCHAR(30),
sobrenome VARCHAR(30),
RG VARCHAR(30),
CPF VARCHAR(30),
data_nascimento VARCHAR(100),
logradouro VARCHAR(100),
complemento VARCHAR(100),
bairro VARCHAR(100),
uf VARCHAR(100),
CEP VARCHAR(100),
dinheiro_real INT(13),
dinheiro_dolar INT(13),
profissao VARCHAR(100),
mercado VARCHAR(100),
salario_real VARCHAR(100),
salario_dolar VARCHAR(100),
lastupdate TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
)

"""



connection = pymysql.connect(host='jobs.visie.com.br',
                             user='daniellomezo',
                             password='0a978b33c438a468ecfabf5645415fec',
                             port=3306,
                             db='daniellomezo')

class Conn_db(object):
    def Consulta(self, sql):
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql)
                result = cursor.fetchall()
                print(result)
        finally:
            connection.close()


    def Insere_dados(self):
        arquivo_excel = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dados-pessoais.xlsx'))
        planilha1 = arquivo_excel['Página1']


        # Começa dados Complementares
        #dinheiro_real = ''
        #dinheiro_dolar = ''
        #profissao = ''
        #mercado = ''
        #salario_real = ''
        #salario_dolar = ''
        for insert in range(1, 51 + 2):
            with open(os.path.join(os.path.dirname(os.path.abspath(__file__)),'dados-complementares.csv')) as arquivo_csv:
                leitor_csv = csv.reader(arquivo_csv, delimiter=',')
                leitor_csv.__next__()
                for row in leitor_csv:
                    dinheiro_real = Busca_moedas.Tratamento_moedas('0', 'Real')
                    dinheiro_dolar = Busca_moedas.Tratamento_moedas('0', 'Real')
                    profissao = Google_Trad.Traduzindo_texto('0', row[1])
                    mercado = Google_Trad.Traduzindo_texto('0', row[2])
                    salario_dolar = row[3].replace('$', '')
                    cambio = Busca_moedas.Tratamento_moedas('0', 'converter')
                    try:
                        salario_real = float(salario_dolar) / float(cambio)
                    except ValueError:
                        salario_real = 0
            for i in range(1, 51 + 2):
                nome_completo = planilha1.cell(row=i, column=1).value
                nome = [nome_completo.strip().split(' ')[0] for nome_completo in nome_completo]
                sobrenome = [nome_completo.strip().split(' ')[0] for nome_completo in nome_completo]
                RG = planilha1.cell(row=i, column=2).value
                CPF = planilha1.cell(row=i, column=3).value
                data_aniversario = planilha1.cell(row=i, column=4).value
                CEP = planilha1.cell(row=i, column=5).value
                logradouro = ''
                complemento = ''
                bairro = ''
                localidade = ''
                uf = ''


                cep = Requesicao_Cep.Pesquisa_Cep('0', planilha1.cell(row=i, column=5).value)
                pprint('Endereço completo {} '.format(cep))
                print(nome, sobrenome, RG, CPF, data_aniversario, CEP, logradouro, complemento, bairro, localidade, uf, dinheiro_real, dinheiro_dolar, profissao, mercado, salario_real, salario_dolar)


Conn_db.Insere_dados('0')



        #sql = "INSERT INTO customers (name, address) VALUES (%s, %s)"
        #val = ("John", "Highway 21")
        #with connection.cursor() as cursor:
        #    cursor.execute(sql, val)
