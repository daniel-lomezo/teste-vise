# -*- coding: utf-8 -*-
from Leitura_dados.Tradutor import Google_Trad
from Leitura_dados.Conversor_moeda import Busca_moedas
from openpyxl import load_workbook
from Leitura_dados.Busca_cep import Requesicao_Cep
import pymysql.cursors
import csv
import os


class Conn_db(object):

    def Consulta(self, sql):
        connection = pymysql.connect(host='jobs.visie.com.br',
                                     user='daniellomezo',
                                     password='0a978b33c438a468ecfabf5645415fec',
                                     port=3306,
                                     db='daniellomezo')
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql)
                result = cursor.fetchall()
                print(result)
        finally:
            connection.close()


    def Insere_dados(self):
        arquivo_excel = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dados-pessoais.xlsx'))
        planilha1 = arquivo_excel['Página1']
        arquivo_csv = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dados-complementares.csv'))
        leitor_csv = csv.reader(arquivo_csv, delimiter=',')
        leitor_csv.__next__()
        i = 2
        j = 1
        for row in leitor_csv:
            try:
                nome, sobrenome = planilha1.cell(row=i, column=1).value.split(" ", 1)
            except AttributeError:
                nome = 'Nome não encontrado'
                sobrenome = 'sobrenome não encontrado'
            RG = planilha1.cell(row=i, column=2).value
            CPF = planilha1.cell(row=i, column=3).value
            data_aniversario = planilha1.cell(row=i, column=4).value
            CEP = planilha1.cell(row=i, column=5).value
            try:
                cep = Requesicao_Cep.Pesquisa_Cep('0', planilha1.cell(row=i, column=5).value)
            except TypeError:
                cep = 'Endereço não encontrado'
            logradouro = cep.get('logradouro')
            complemento = cep.get('complemento')
            bairro = cep.get('bairro')
            localidade = cep.get('localidade')
            uf = cep.get('uf')
            """
            logradouro = 'Não Encontrado'
            complemento = 'Não Encontrado'
            bairro = 'Não Encontrado'
            localidade = 'Não Encontrado'
            uf = 'Não Encontrado'"""
            dinheiro_real = Busca_moedas.Tratamento_moedas('0', 'Real')
            dinheiro_dolar = Busca_moedas.Tratamento_moedas('0', 'Dolar')
            profissao = Google_Trad.Traduzindo_texto('0', row[1])
            mercado = Google_Trad.Traduzindo_texto('0', row[2])
            salario_dolar = row[3].replace('$', '')
            cambio = Busca_moedas.Tratamento_moedas('0', 'converter')
            try:
                salario_real = float(salario_dolar) / float(cambio)
            except ValueError:
                salario_real = 0
            val = nome, sobrenome, RG, CPF, data_aniversario, CEP, logradouro, complemento, bairro, localidade, uf, dinheiro_real, dinheiro_dolar, profissao, mercado, salario_real, salario_dolar
            sql = """INSERT INTO Desafio (nome, sobrenome, RG, CPF, 
                                            data_nascimento, CEP, logradouro, 
                                            complemento, bairro, localidade, 
                                            uf, dinheiro_real, dinheiro_dolar, 
                                            profissao, mercado, salario_real, 
                                            salario_dolar) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

            connection = pymysql.connect(host='jobs.visie.com.br',
                                         user='daniellomezo',
                                         password='0a978b33c438a468ecfabf5645415fec',
                                         port=3306,
                                         db='daniellomezo')
            with connection.cursor() as cursor:
             cursor.execute(sql, val)
            connection.commit()

            i += 1
            j += 1
            print('Indice em: ', i)
            print('')
            print('')




